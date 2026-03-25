import numpy as np
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime
import os
import requests
from bs4 import BeautifulSoup

st.set_page_config(page_title="Trading Dashboard", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;600&display=swap');

    :root {
        --blue-900: #0a1628;
        --blue-800: #0f2441;
        --blue-700: #14365a;
        --blue-600: #1a4f8b;
        --blue-500: #2563eb;
        --blue-400: #3b82f6;
        --blue-300: #60a5fa;
        --blue-200: #93c5fd;
        --blue-100: #dbeafe;
        --blue-50:  #eff6ff;
        --accent:   #0ea5e9;
        --accent-2: #06b6d4;
        --success:  #10b981;
        --danger:   #ef4444;
        --warning:  #f59e0b;
        --surface:  #ffffff;
        --surface-2:#f8fafc;
        --surface-3:#f1f5f9;
        --border:   #e2e8f0;
        --border-2: #cbd5e1;
        --text:     #0f172a;
        --text-2:   #334155;
        --text-3:   #64748b;
        --text-4:   #94a3b8;
        --card-shadow: 0 2px 4px rgba(0,0,0,0.04), 0 8px 20px rgba(15,23,42,0.08), 0 1px 2px rgba(0,0,0,0.06);
        --card-shadow-hover: 0 8px 16px rgba(0,0,0,0.08), 0 20px 40px rgba(15,23,42,0.12), 0 2px 4px rgba(0,0,0,0.06);
    }

    .stApp {
        background: #eef2f7 !important;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
    }

    #MainMenu, footer, .stDeployButton { display: none !important; }
    div[data-testid="stToolbar"] { display: none !important; }
    header[data-testid="stHeader"] { display: none !important; }

    /* Ocultar botón cerrar sidebar — siempre visible */
    section[data-testid="stSidebar"] button[kind="headerNoPadding"],
    section[data-testid="stSidebar"] [data-testid="stBaseButton-headerNoPadding"],
    section[data-testid="stSidebar"] > div:first-child > div:first-child > div > button {
        display: none !important;
    }

    /* ═══════════════════════════════════════════════
       SIDEBAR STYLING
       ═══════════════════════════════════════════════ */
    section[data-testid="stSidebar"] {
        background: linear-gradient(195deg, var(--blue-900) 0%, var(--blue-800) 40%, #0c1e3a 100%) !important;
        border-right: 1px solid rgba(37, 99, 235, 0.15) !important;
        box-shadow: 4px 0 30px rgba(10, 22, 40, 0.3) !important;
    }
    section[data-testid="stSidebar"] * {
        color: #e2e8f0 !important;
        font-family: 'Inter', sans-serif !important;
    }
    section[data-testid="stSidebar"] .stTextInput input,
    section[data-testid="stSidebar"] .stNumberInput input,
    section[data-testid="stSidebar"] .stSelectbox select,
    section[data-testid="stSidebar"] .stDateInput input {
        background: rgba(255,255,255,0.06) !important;
        border: 1px solid rgba(59, 130, 246, 0.25) !important;
        border-radius: 10px !important;
        color: #f1f5f9 !important;
        font-family: 'Inter', sans-serif !important;
        transition: all 0.3s ease !important;
    }
    section[data-testid="stSidebar"] .stTextInput input:focus,
    section[data-testid="stSidebar"] .stNumberInput input:focus {
        border-color: var(--blue-400) !important;
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.15) !important;
    }
    section[data-testid="stSidebar"] .stButton > button {
        background: linear-gradient(135deg, var(--blue-500), var(--accent)) !important;
        color: white !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 12px 24px !important;
        font-weight: 700 !important;
        font-size: 14px !important;
        letter-spacing: 0.02em !important;
        transition: all 0.4s cubic-bezier(0.23, 1, 0.32, 1) !important;
        box-shadow: 0 4px 15px rgba(37, 99, 235, 0.3) !important;
        width: 100% !important;
    }
    section[data-testid="stSidebar"] .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 25px rgba(37, 99, 235, 0.45) !important;
    }
    section[data-testid="stSidebar"] hr {
        border-color: rgba(59, 130, 246, 0.15) !important;
    }

    /* ═══════════════════════════════════════════════
       MAIN HEADER
       ═══════════════════════════════════════════════ */
    .main-header {
        background: linear-gradient(135deg, var(--blue-900) 0%, var(--blue-700) 50%, var(--blue-600) 100%);
        border-radius: 20px;
        padding: 40px 48px;
        margin-bottom: 32px;
        position: relative;
        overflow: hidden;
        box-shadow: 0 20px 60px rgba(10, 22, 40, 0.25);
        animation: headerSlide 0.8s cubic-bezier(0.23, 1, 0.32, 1);
    }
    .main-header::before {
        content: '';
        position: absolute;
        top: -50%; right: -20%;
        width: 500px; height: 500px;
        background: radial-gradient(circle, rgba(59,130,246,0.12) 0%, transparent 70%);
        border-radius: 50%;
    }
    .main-header::after {
        content: '';
        position: absolute;
        bottom: -30%; left: 10%;
        width: 300px; height: 300px;
        background: radial-gradient(circle, rgba(14,165,233,0.08) 0%, transparent 70%);
        border-radius: 50%;
    }
    .main-header h1 {
        font-family: 'Inter', sans-serif !important;
        font-size: 36px !important;
        font-weight: 900 !important;
        color: #ffffff !important;
        margin: 0 !important;
        letter-spacing: -0.03em;
        position: relative; z-index: 1;
    }
    .main-header p {
        color: var(--blue-200) !important;
        font-size: 14px; margin-top: 8px;
        font-weight: 500; letter-spacing: 0.02em;
        position: relative; z-index: 1;
    }
    .header-badge {
        display: inline-flex; align-items: center; gap: 6px;
        background: rgba(255,255,255,0.08);
        backdrop-filter: blur(10px);
        border: 1px solid rgba(255,255,255,0.12);
        border-radius: 100px; padding: 6px 16px;
        font-size: 10px; font-weight: 800;
        color: var(--blue-200);
        letter-spacing: 0.12em; text-transform: uppercase;
        margin-bottom: 14px;
        position: relative; z-index: 1;
    }
    .header-badge .pulse {
        width: 7px; height: 7px;
        background: var(--success);
        border-radius: 50%;
        animation: pulse 2s infinite;
    }
    @keyframes pulse {
        0%, 100% { opacity: 1; transform: scale(1); }
        50% { opacity: 0.5; transform: scale(1.4); }
    }
    @keyframes headerSlide {
        from { opacity: 0; transform: translateY(-20px); }
        to { opacity: 1; transform: translateY(0); }
    }

    /* ═══════════════════════════════════════════════
       KPI CARDS
       ═══════════════════════════════════════════════ */
    .kpi-grid {
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 16px; margin-bottom: 24px;
    }
    .kpi-card {
        background: #ffffff;
        border: 1px solid rgba(226, 232, 240, 0.7);
        border-radius: 16px; padding: 24px;
        position: relative; overflow: hidden;
        transition: all 0.4s cubic-bezier(0.23, 1, 0.32, 1);
        animation: cardFadeIn 0.6s ease forwards;
        opacity: 0;
        box-shadow: var(--card-shadow);
    }
    .kpi-card:hover {
        transform: translateY(-4px);
        box-shadow: var(--card-shadow-hover);
    }
    .kpi-card::before {
        content: '';
        position: absolute; top: 0; left: 0; right: 0;
        height: 3px;
        background: linear-gradient(90deg, var(--blue-500), var(--accent));
        opacity: 0; transition: opacity 0.3s ease;
    }
    .kpi-card:hover::before { opacity: 1; }
    .kpi-card:nth-child(1) { animation-delay: 0.05s; }
    .kpi-card:nth-child(2) { animation-delay: 0.1s; }
    .kpi-card:nth-child(3) { animation-delay: 0.15s; }
    .kpi-card:nth-child(4) { animation-delay: 0.2s; }
    .kpi-card:nth-child(5) { animation-delay: 0.25s; }
    .kpi-card:nth-child(6) { animation-delay: 0.3s; }
    .kpi-card:nth-child(7) { animation-delay: 0.35s; }
    .kpi-card:nth-child(8) { animation-delay: 0.4s; }
    @keyframes cardFadeIn {
        from { opacity: 0; transform: translateY(16px); }
        to { opacity: 1; transform: translateY(0); }
    }

    .kpi-icon { width: 40px; height: 40px; border-radius: 12px; display: flex; align-items: center; justify-content: center; font-size: 18px; margin-bottom: 14px; }
    .kpi-icon.blue { background: var(--blue-50); }
    .kpi-icon.green { background: #ecfdf5; }
    .kpi-icon.red { background: #fef2f2; }
    .kpi-icon.amber { background: #fffbeb; }

    .kpi-label { font-size: 11px; font-weight: 700; color: var(--text-4); text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 6px; }
    .kpi-value { font-family: 'Inter', sans-serif; font-size: 28px; font-weight: 800; color: var(--text); line-height: 1; letter-spacing: -0.02em; }
    .kpi-value.accent { color: var(--blue-500); }
    .kpi-value.success { color: var(--success); }
    .kpi-value.danger { color: var(--danger); }
    .kpi-sub { font-size: 11px; color: var(--text-3); margin-top: 6px; font-weight: 500; }

    .estado-badge { display: inline-flex; align-items: center; gap: 10px; padding: 14px 28px; border-radius: 14px; font-weight: 700; font-size: 15px; letter-spacing: 0.01em; animation: cardFadeIn 0.6s ease 0.5s forwards; opacity: 0; }
    .estado-badge.rentable { background: linear-gradient(135deg, #ecfdf5, #d1fae5); color: #065f46; border: 1px solid #a7f3d0; box-shadow: 0 4px 12px rgba(16, 185, 129, 0.15); }
    .estado-badge.warning { background: linear-gradient(135deg, #fffbeb, #fef3c7); color: #92400e; border: 1px solid #fde68a; box-shadow: 0 4px 12px rgba(245, 158, 11, 0.15); }
    .estado-badge.danger { background: linear-gradient(135deg, #fef2f2, #fee2e2); color: #991b1b; border: 1px solid #fecaca; box-shadow: 0 4px 12px rgba(239, 68, 68, 0.15); }
    .estado-badge.neutral { background: linear-gradient(135deg, var(--surface-2), var(--surface-3)); color: var(--text-2); border: 1px solid var(--border); box-shadow: var(--card-shadow); }

    .section-title { font-family: 'Inter', sans-serif; font-size: 22px; font-weight: 800; color: var(--text); margin: 40px 0 8px 0; letter-spacing: -0.02em; animation: cardFadeIn 0.6s ease forwards; opacity: 0; }
    .section-sub { font-size: 13px; color: var(--text-3); margin-bottom: 24px; font-weight: 500; }

    /* ═══════════════════════════════════════════════
       PLOTLY CHARTS — shadow directo al elemento real
       ═══════════════════════════════════════════════ */
    div[data-testid="stPlotlyChart"] {
        background: #ffffff !important;
        border: 1px solid rgba(226, 232, 240, 0.7) !important;
        border-radius: 16px !important;
        padding: 16px 12px 8px 12px !important;
        box-shadow: var(--card-shadow) !important;
        transition: all 0.4s ease !important;
        margin-bottom: 12px !important;
    }
    div[data-testid="stPlotlyChart"]:hover {
        box-shadow: var(--card-shadow-hover) !important;
    }

    .js-plotly-plot .plotly .modebar { display: none !important; }
    div[data-testid="stMetric"] { display: none !important; }

    /* ═══════════════════════════════════════════════
       FILE UPLOADER
       ═══════════════════════════════════════════════ */
    div[data-testid="stFileUploader"] { border: 2px dashed #cbd5e1 !important; border-radius: 16px !important; padding: 20px !important; background: #ffffff !important; transition: all 0.3s ease !important; box-shadow: var(--card-shadow); }
    div[data-testid="stFileUploader"]:hover { border-color: var(--blue-300) !important; background: var(--blue-50) !important; }

    /* ═══════════════════════════════════════════════
       EXPANDER — "Ver todas las operaciones"
       ═══════════════════════════════════════════════ */
    div[data-testid="stExpander"] {
        background: #ffffff !important;
        border: 1px solid rgba(226, 232, 240, 0.7) !important;
        border-radius: 16px !important;
        box-shadow: var(--card-shadow) !important;
        overflow: hidden !important;
        transition: all 0.3s ease !important;
    }
    div[data-testid="stExpander"]:hover {
        box-shadow: var(--card-shadow-hover) !important;
    }
    div[data-testid="stExpander"] summary {
        background: #ffffff !important;
        border: none !important;
        border-radius: 16px !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 700 !important;
        font-size: 14px !important;
        color: var(--text) !important;
        padding: 18px 24px !important;
        transition: background 0.2s ease !important;
    }
    div[data-testid="stExpander"] summary:hover {
        background: var(--surface-2) !important;
    }
    div[data-testid="stExpander"] summary span {
        color: var(--text) !important;
        font-weight: 700 !important;
    }
    div[data-testid="stExpander"] summary svg {
        color: var(--blue-500) !important;
    }
    div[data-testid="stExpander"] div[data-testid="stExpanderDetails"] {
        padding: 0 24px 20px 24px !important;
        border-top: 1px solid var(--border) !important;
    }

    /* ═══════════════════════════════════════════════
       DATAFRAME INSIDE EXPANDER
       ═══════════════════════════════════════════════ */
    div[data-testid="stDataFrame"] {
        border-radius: 12px !important;
        overflow: hidden !important;
    }

    /* ═══════════════════════════════════════════════
       DOWNLOAD BUTTONS
       ═══════════════════════════════════════════════ */
    .stDownloadButton > button {
        background: #ffffff !important;
        border: 1px solid rgba(226, 232, 240, 0.7) !important;
        border-radius: 12px !important;
        color: var(--text) !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 700 !important;
        padding: 12px 28px !important;
        transition: all 0.3s ease !important;
        box-shadow: var(--card-shadow);
    }
    .stDownloadButton > button:hover {
        color: var(--blue-500) !important;
        box-shadow: var(--card-shadow-hover) !important;
        transform: translateY(-2px) !important;
    }

    .stSuccess, .stWarning, .stError, .stInfo { border-radius: 14px !important; font-family: 'Inter', sans-serif !important; border: none !important; }

    .balance-warning { background: linear-gradient(135deg, #fffbeb, #fef3c7); border: 1px solid #fde68a; border-radius: 12px; padding: 14px 20px; font-size: 13px; color: #92400e; font-weight: 600; margin-bottom: 16px; display: flex; align-items: center; gap: 10px; box-shadow: var(--card-shadow); }

    /* ═══════════════════════════════════════════════
       NOTICIAS
       ═══════════════════════════════════════════════ */
    .no-news-card {
        background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
        border: 1px solid rgba(96, 165, 250, 0.3);
        border-radius: 16px;
        padding: 44px 32px;
        text-align: center;
        animation: cardFadeIn 0.7s ease forwards;
        opacity: 0;
        box-shadow: var(--card-shadow);
    }
    .no-news-icon { font-size: 48px; margin-bottom: 14px; line-height: 1; }
    .no-news-title { font-family: 'Inter', sans-serif; font-size: 17px; font-weight: 800; color: var(--blue-600); margin-bottom: 8px; letter-spacing: -0.01em; }
    .no-news-sub { font-size: 13px; color: var(--text-3); font-weight: 500; max-width: 400px; margin: 0 auto; line-height: 1.6; }

    .news-item {
        background: #ffffff;
        border: 1px solid rgba(226, 232, 240, 0.7);
        border-radius: 14px;
        padding: 16px 20px;
        margin-bottom: 8px;
        box-shadow: var(--card-shadow);
        transition: all 0.3s ease;
        animation: cardFadeIn 0.6s ease forwards;
        opacity: 0;
    }
    .news-item:hover {
        box-shadow: var(--card-shadow-hover);
        transform: translateY(-2px);
    }

    /* ═══════════════════════════════════════════════
       DIVIDER
       ═══════════════════════════════════════════════ */
    hr { border-color: rgba(226, 232, 240, 0.5) !important; }

    @media (max-width: 768px) { .kpi-grid { grid-template-columns: repeat(2, 1fr); } .main-header { padding: 28px 24px; } .main-header h1 { font-size: 26px !important; } }
</style>
""", unsafe_allow_html=True)


def chart_layout(title="", xaxis_title="", yaxis_title="", height=400, showlegend=False):
    return dict(
        title=dict(text=f"<b>{title}</b>", font=dict(family="Inter", size=15, color="#0f172a"), x=0.02, y=0.96),
        xaxis=dict(title=dict(text=xaxis_title, font=dict(family="Inter", size=12, color="#64748b")), gridcolor="#f1f5f9", linecolor="#e2e8f0", tickfont=dict(family="Inter", size=11, color="#94a3b8"), zeroline=False),
        yaxis=dict(title=dict(text=yaxis_title, font=dict(family="Inter", size=12, color="#64748b")), gridcolor="#f1f5f9", linecolor="rgba(0,0,0,0)", tickfont=dict(family="Inter", size=11, color="#94a3b8"), zeroline=True, zerolinecolor="#e2e8f0", zerolinewidth=1),
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", height=height,
        margin=dict(l=60, r=30, t=60, b=50), showlegend=showlegend,
        legend=dict(font=dict(family="Inter", size=12, color="#334155"), bgcolor="rgba(255,255,255,0.8)", bordercolor="#e2e8f0", borderwidth=1),
        hoverlabel=dict(bgcolor="#0f172a", font=dict(family="Inter", size=12, color="white"), bordercolor="rgba(0,0,0,0)"),
    )


CSV_PATH = "operaciones_trading.csv"

if "data" not in st.session_state:
    if os.path.exists(CSV_PATH):
        st.session_state.data = pd.read_csv(CSV_PATH)
    else:
        st.session_state.data = pd.DataFrame(columns=["Fecha", "Par", "Tipo", "SL (pips)", "TP (pips)", "Resultado USD"])

if "balance_inicial" not in st.session_state:
    st.session_state.balance_inicial = 0

df = st.session_state.data.copy()
df = df.drop_duplicates()
df["Fecha"] = pd.to_datetime(df["Fecha"], errors='coerce')
df = df.dropna(subset=["Fecha"])
df = df.sort_values("Fecha").reset_index(drop=True)

if "Resultado USD" in df.columns:
    df["Resultado USD"] = pd.to_numeric(df["Resultado USD"], errors="coerce")
else:
    st.error("La columna 'Resultado USD' no existe en los datos.")
    st.stop()

df["Resultado"] = df["Resultado USD"].apply(lambda x: "Win" if x > 0 else ("Loss" if x < 0 else "BreakEven"))
df_operaciones = df.copy()

with st.sidebar:
    st.markdown("""
    <div style='text-align:center; padding: 20px 0 10px 0;'>
        <div style='font-family: Inter, sans-serif; font-size: 22px; font-weight: 900; color: #fff; letter-spacing: -0.03em;'>Trading</div>
        <div style='font-family: Inter, sans-serif; font-size: 10px; color: #60a5fa; letter-spacing: 0.18em; text-transform: uppercase; font-weight: 700;'>Dashboard Pro</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("<p style='font-size:10px; text-transform:uppercase; letter-spacing:0.12em; color:#60a5fa; font-weight:700; margin-bottom:4px;'>⚙️ Parámetros</p>", unsafe_allow_html=True)
    st.session_state.balance_inicial = st.number_input(
        "Balance TOTAL de la cuenta ($)", value=int(st.session_state.balance_inicial), step=100, format="%d",
        help="Ingresa el balance TOTAL de tu cuenta (ej: 1000, 5000, 10000). Los % se calculan sobre este valor."
    )
    st.markdown("---")

    st.markdown("<p style='font-size:10px; text-transform:uppercase; letter-spacing:0.12em; color:#60a5fa; font-weight:700; margin-bottom:4px;'>📥 Nueva operación</p>", unsafe_allow_html=True)
    fecha = st.date_input("Fecha", value=datetime.today())

    # ── Lista completa de pares Forex + Commodities + Índices ──
    PARES_FOREX = {
        "💱 Majors": [
            "EURUSD", "GBPUSD", "USDJPY", "USDCHF", "AUDUSD", "USDCAD", "NZDUSD"
        ],
        "🔄 Crosses EUR": [
            "EURGBP", "EURJPY", "EURCHF", "EURAUD", "EURCAD", "EURNZD"
        ],
        "🔄 Crosses GBP": [
            "GBPJPY", "GBPCHF", "GBPAUD", "GBPCAD", "GBPNZD"
        ],
        "🔄 Crosses JPY": [
            "AUDJPY", "CADJPY", "CHFJPY", "NZDJPY"
        ],
        "🔄 Crosses AUD/NZD/CAD": [
            "AUDCAD", "AUDCHF", "AUDNZD", "NZDCAD", "NZDCHF", "CADCHF"
        ],
        "🥇 Metales": [
            "XAUUSD", "XAGUSD", "XAUEUR", "XAUGBP", "XAUJPY",
            "XPTUSD", "XPDUSD"
        ],
        "🛢️ Energía": [
            "USOIL", "UKOIL", "NGAS"
        ],
        "📊 Índices": [
            "US30", "US100", "US500", "DE40", "UK100", "JP225",
            "FR40", "AU200", "HK50", "USTEC", "VIX"
        ],
        "🪙 Crypto CFDs": [
            "BTCUSD", "ETHUSD", "LTCUSD", "XRPUSD", "SOLUSD"
        ],
        "💵 Exóticos": [
            "USDMXN", "USDZAR", "USDTRY", "USDSGD", "USDHKD",
            "USDSEK", "USDNOK", "USDDKK", "USDPLN", "USDCZK",
            "USDHUF", "EURTRY", "EURMXN", "EURZAR", "EURSEK",
            "EURNOK", "EURPLN", "EURHUF", "GBPZAR", "GBPMXN",
            "GBPSEK", "GBPNOK", "GBPTRY"
        ],
    }

    # Construir lista plana con separadores de categoría
    opciones_pares = []
    for categoria, pares in PARES_FOREX.items():
        opciones_pares.append(f"── {categoria} ──")
        opciones_pares.extend(pares)

    par_seleccionado = st.selectbox(
        "Par / Instrumento",
        options=opciones_pares,
        index=1,  # Default: EURUSD
        help="Selecciona el par o instrumento operado"
    )
    # Si eligieron un separador de categoría, forzar EURUSD
    par = par_seleccionado if not par_seleccionado.startswith("──") else "EURUSD"

    tipo = st.selectbox("Tipo", ["Compra", "Venta"])
    resultado_usd = st.number_input("Resultado USD", step=1, format="%d")
    sl = st.number_input("SL (pips)", min_value=0, step=1, format="%d")
    tp = st.number_input("TP (pips)", min_value=0, step=1, format="%d")
    if st.button("➕  Agregar operación"):
        nueva_op = pd.DataFrame([{"Fecha": fecha, "Par": par, "Tipo": tipo, "SL (pips)": sl, "TP (pips)": tp, "Resultado USD": resultado_usd}])
        st.session_state.data = pd.concat([st.session_state.data, nueva_op], ignore_index=True)
        st.session_state.data.to_csv(CSV_PATH, index=False)
        st.success("✓ Operación agregada")

    st.markdown("---")
    st.markdown("<p style='font-size:10px; text-transform:uppercase; letter-spacing:0.12em; color:#ef4444; font-weight:700; margin-bottom:4px;'>🗑️ Zona de riesgo</p>", unsafe_allow_html=True)
    confirmar_reset = st.checkbox("Confirmar que quiero borrar TODO", value=False)
    if st.button("🗑️  Resetear datos", disabled=not confirmar_reset):
        st.session_state.data = pd.DataFrame(columns=["Fecha", "Par", "Tipo", "SL (pips)", "TP (pips)", "Resultado USD"])
        if os.path.exists(CSV_PATH):
            os.remove(CSV_PATH)
        st.session_state.balance_inicial = 0
        st.success("✓ Datos reseteados")
        st.rerun()

st.markdown("""
<div style='margin-bottom: 8px;'>
    <span style='font-family: Inter; font-size: 11px; font-weight: 700; color: #64748b; text-transform: uppercase; letter-spacing: 0.1em;'>📂 Cargar archivo CSV externo</span>
</div>
""", unsafe_allow_html=True)
csv_file = st.file_uploader("", type="csv", label_visibility="collapsed")
if csv_file is not None:
    uploaded_data = pd.read_csv(csv_file)
    st.session_state.data = pd.concat([st.session_state.data, uploaded_data], ignore_index=True)
    st.session_state.data.to_csv(CSV_PATH, index=False)
    st.success("Archivo cargado y datos añadidos")

df = st.session_state.data.copy()
df["Fecha"] = pd.to_datetime(df["Fecha"], errors='coerce')
df = df.dropna(subset=["Fecha"])
df = df.sort_values("Fecha").reset_index(drop=True)

if not df.empty and df.iloc[0]["Tipo"] != "Inicial":
    fila_inicial = pd.DataFrame([{"Fecha": df["Fecha"].min() - pd.Timedelta(days=1), "Par": "INICIAL", "Tipo": "Inicial", "SL (pips)": 0, "TP (pips)": 0, "Resultado USD": 0, "Resultado": "Neutral"}])
    df = pd.concat([fila_inicial, df], ignore_index=True)

df["Resultado"] = df["Resultado USD"].apply(lambda x: "Win" if x > 0 else "Loss")
df["Balance Acumulado"] = st.session_state.balance_inicial + df["Resultado USD"].cumsum()
df["Operación #"] = range(1, len(df) + 1)
df_visual = df.copy()
df_operaciones = df[df["Tipo"] != "Inicial"].copy()

st.markdown("""
<div class="main-header">
    <div class="header-badge"><div class="pulse"></div>LIVE TRACKING</div>
    <h1>Trading Dashboard</h1>
    <p>Análisis inteligente de rendimiento · Métricas en tiempo real · Predicción de tendencia</p>
</div>
""", unsafe_allow_html=True)

total_ops = len(df_operaciones)
wins = (df_operaciones["Resultado USD"] > 0).sum()
losses = total_ops - wins
winrate = wins / total_ops if total_ops > 0 else 0

ganancias = df_operaciones[df_operaciones["Resultado USD"] > 0]["Resultado USD"].sum()
perdidas = df_operaciones[df_operaciones["Resultado USD"] < 0]["Resultado USD"].sum()
ganancia_neta = ganancias + perdidas
balance_inicial = st.session_state.balance_inicial
balance_final = df["Balance Acumulado"].iloc[-1] if not df.empty else balance_inicial

rr_promedio = (df_operaciones["TP (pips)"] / df_operaciones["SL (pips)"]).replace([float("inf"), -float("inf")], 0).fillna(0).mean()
rentabilidad_esperada = (winrate * rr_promedio) - (1 - winrate)
winrate_minimo = 1 / (1 + rr_promedio) if rr_promedio > 0 else 1.0

estado_rentabilidad = "🔍 Sin suficiente data"
estado_class = "neutral"
if total_ops >= 3:
    perdidas_relativas = (balance_inicial - balance_final) / balance_inicial if balance_inicial > 0 else 0
    if perdidas_relativas >= 0.10:
        estado_rentabilidad = "❌ Quemó la cuenta"; estado_class = "danger"
    elif winrate < 0.30:
        estado_rentabilidad = "❌ No rentable — Winrate muy bajo"; estado_class = "danger"
    elif rentabilidad_esperada > 0 and ganancia_neta > 0:
        estado_rentabilidad = "✅ Rentable"; estado_class = "rentable"
    elif rentabilidad_esperada > 0 and ganancia_neta <= 0:
        estado_rentabilidad = "⚠️ Rentabilidad potencial (en pérdida)"; estado_class = "warning"
    else:
        estado_rentabilidad = "❌ No rentable"; estado_class = "danger"

ganancia_color = "success" if ganancia_neta >= 0 else "danger"
balance_color = "accent" if balance_final >= balance_inicial else "danger"

# ── Cálculos avanzados (se usan junto con los generales) ──
profit_factor = abs(ganancias / perdidas) if perdidas != 0 else float('inf')
profit_factor_display = f"{profit_factor:.2f}" if profit_factor != float('inf') else "∞"
pf_color = "success" if profit_factor >= 1.5 else ("accent" if profit_factor >= 1.0 else "danger")
pf_icon_color = "green" if profit_factor >= 1.5 else ("blue" if profit_factor >= 1.0 else "red")

balances = df["Balance Acumulado"].values
peak = np.maximum.accumulate(balances)
drawdown = (peak - balances)
max_dd = drawdown.max() if len(drawdown) > 0 else 0
max_dd_pct = (max_dd / peak[np.argmax(drawdown)] * 100) if len(drawdown) > 0 and peak[np.argmax(drawdown)] > 0 else 0

avg_win = df_operaciones[df_operaciones["Resultado USD"] > 0]["Resultado USD"].mean() if wins > 0 else 0
avg_loss = abs(df_operaciones[df_operaciones["Resultado USD"] < 0]["Resultado USD"].mean()) if losses > 0 else 0

rachas_win = 0
rachas_loss = 0
racha_actual_w = 0
racha_actual_l = 0
for r in df_operaciones["Resultado USD"]:
    if r > 0:
        racha_actual_w += 1
        racha_actual_l = 0
    else:
        racha_actual_l += 1
        racha_actual_w = 0
    rachas_win = max(rachas_win, racha_actual_w)
    rachas_loss = max(rachas_loss, racha_actual_l)

roi = ((balance_final - balance_inicial) / balance_inicial * 100) if balance_inicial > 0 else 0
roi_color = "success" if roi >= 0 else "danger"
roi_icon = "green" if roi >= 0 else "red"
expectancy = avg_win * winrate - avg_loss * (1 - winrate)

st.markdown(f"""
<div class="section-title">Estadísticas Generales</div>
<div class="section-sub">Resumen completo de rendimiento, riesgo y consistencia</div>
<div class="kpi-grid">
    <div class="kpi-card"><div class="kpi-icon blue">📊</div><div class="kpi-label">Total operaciones</div><div class="kpi-value">{total_ops}</div><div class="kpi-sub">{wins} ganadas · {losses} perdidas</div></div>
    <div class="kpi-card"><div class="kpi-icon blue">🎯</div><div class="kpi-label">Winrate</div><div class="kpi-value accent">{winrate * 100:.1f}%</div><div class="kpi-sub">Mínimo requerido: {winrate_minimo * 100:.1f}%</div></div>
    <div class="kpi-card"><div class="kpi-icon {'green' if ganancia_neta >= 0 else 'red'}">💵</div><div class="kpi-label">Ganancia neta</div><div class="kpi-value {ganancia_color}">${ganancia_neta:,.0f}</div><div class="kpi-sub">Ganancias ${ganancias:,.0f} · Pérdidas ${perdidas:,.0f}</div></div>
    <div class="kpi-card"><div class="kpi-icon {pf_icon_color}">🏆</div><div class="kpi-label">Profit Factor</div><div class="kpi-value {pf_color}">{profit_factor_display}</div><div class="kpi-sub">{'Excelente' if profit_factor >= 2 else ('Bueno' if profit_factor >= 1.5 else ('Ajustado' if profit_factor >= 1 else 'Negativo'))}</div></div>
    <div class="kpi-card"><div class="kpi-icon blue">💼</div><div class="kpi-label">Balance inicial</div><div class="kpi-value">${balance_inicial:,.0f}</div><div class="kpi-sub">Capital de partida</div></div>
    <div class="kpi-card"><div class="kpi-icon {'green' if balance_final >= balance_inicial else 'red'}">💰</div><div class="kpi-label">Balance final</div><div class="kpi-value {balance_color}">${balance_final:,.0f}</div><div class="kpi-sub">{'↑' if balance_final >= balance_inicial else '↓'} ${abs(balance_final - balance_inicial):,.0f} vs inicial</div></div>
    <div class="kpi-card"><div class="kpi-icon blue">⚖️</div><div class="kpi-label">RR Promedio</div><div class="kpi-value">{rr_promedio:.2f}</div><div class="kpi-sub">Ratio riesgo/beneficio</div></div>
    <div class="kpi-card"><div class="kpi-icon {roi_icon}">💹</div><div class="kpi-label">ROI</div><div class="kpi-value {roi_color}">{roi:+.2f}%</div><div class="kpi-sub">Retorno sobre capital inicial</div></div>
    <div class="kpi-card"><div class="kpi-icon green">✅</div><div class="kpi-label">Promedio Win</div><div class="kpi-value success">${avg_win:,.2f}</div><div class="kpi-sub">Ganancia promedio por trade ganador</div></div>
    <div class="kpi-card"><div class="kpi-icon red">❌</div><div class="kpi-label">Promedio Loss</div><div class="kpi-value danger">${avg_loss:,.2f}</div><div class="kpi-sub">Pérdida promedio por trade perdedor</div></div>
    <div class="kpi-card"><div class="kpi-icon red">📉</div><div class="kpi-label">Máx Drawdown</div><div class="kpi-value danger">${max_dd:,.0f}</div><div class="kpi-sub">{max_dd_pct:.1f}% desde el pico máximo</div></div>
    <div class="kpi-card"><div class="kpi-icon blue">🧮</div><div class="kpi-label">Expectancy</div><div class="kpi-value accent">${expectancy:+,.2f}</div><div class="kpi-sub">Ganancia esperada por operación</div></div>
    <div class="kpi-card"><div class="kpi-icon {'green' if rentabilidad_esperada > 0 else 'red'}">📈</div><div class="kpi-label">Rentabilidad esperada</div><div class="kpi-value {'success' if rentabilidad_esperada > 0 else 'danger'}">{rentabilidad_esperada:.2f}</div><div class="kpi-sub">{'Positiva' if rentabilidad_esperada > 0 else 'Negativa'}</div></div>
    <div class="kpi-card"><div class="kpi-icon green">🔥</div><div class="kpi-label">Racha máx Wins</div><div class="kpi-value success">{rachas_win}</div><div class="kpi-sub">Wins consecutivos</div></div>
    <div class="kpi-card"><div class="kpi-icon red">❄️</div><div class="kpi-label">Racha máx Losses</div><div class="kpi-value danger">{rachas_loss}</div><div class="kpi-sub">Losses consecutivos</div></div>
    <div class="kpi-card"><div class="kpi-icon amber">🔻</div><div class="kpi-label">Winrate mínimo</div><div class="kpi-value">{winrate_minimo * 100:.1f}%</div><div class="kpi-sub">Para ser rentable con tu RR</div></div>
</div>
<div class="estado-badge {estado_class}">{estado_rentabilidad}</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
#  ANÁLISIS VISUAL — gráficos (sin div wrappers manuales)
# ═══════════════════════════════════════════════════════════════
st.markdown('<div class="section-title" style="animation-delay:0.3s;">Análisis Visual</div>', unsafe_allow_html=True)
st.markdown('<div class="section-sub">Gráficos interactivos de tu rendimiento</div>', unsafe_allow_html=True)

colA, colB = st.columns(2)

with colA:
    df_vis = df[df["Tipo"] != "Inicial"].copy()
    if not df_vis.empty:
        # P&L acumulado desde 0 (cada barra muestra ganancia/pérdida neta)
        pnl_acumulado = df_vis["Resultado USD"].cumsum().tolist()
        ops_nums = df_vis["Operación #"].tolist()
        # Insertar punto inicial en 0
        pnl_acumulado = [0] + pnl_acumulado
        ops_nums = [0] + ops_nums
    else:
        pnl_acumulado = [0]
        ops_nums = [0]
    # Colores: verde si P&L >= 0, rojo si < 0
    line_color = "#2563eb" if pnl_acumulado[-1] >= 0 else "#ef4444"
    fill_color = "rgba(37, 99, 235, 0.06)" if pnl_acumulado[-1] >= 0 else "rgba(239, 68, 68, 0.06)"
    fig_balance = go.Figure()
    fig_balance.add_trace(go.Scatter(
        x=ops_nums, y=pnl_acumulado,
        mode="lines+markers", fill="tozeroy", fillcolor=fill_color,
        line=dict(color=line_color, width=3, shape="spline"),
        marker=dict(size=7, color=line_color, line=dict(width=2, color="#fff")),
        hovertemplate="<b>Op #%{x}</b><br>P&L: $%{y:+,.0f}<extra></extra>"
    ))
    # Línea de referencia en 0
    fig_balance.add_hline(y=0, line_dash="dot", line_color="#94a3b8", line_width=1)
    fig_balance.update_layout(**chart_layout("P&L Acumulado", "Operación N°", "USD"))
    fig_balance.update_layout(xaxis=dict(tickmode='linear', tick0=0, dtick=1))
    st.plotly_chart(fig_balance, use_container_width=True)

with colB:
    pares_count = df[df["Tipo"] != "Inicial"]["Par"].value_counts()
    colors_pie = ["#2563eb", "#0ea5e9", "#06b6d4", "#60a5fa", "#93c5fd", "#38bdf8", "#7dd3fc", "#a5f3fc"]
    fig_par = go.Figure(data=[go.Pie(labels=pares_count.index, values=pares_count.values, hole=0.55, marker=dict(colors=colors_pie[:len(pares_count)], line=dict(color="#fff", width=3)), textfont=dict(family="Inter", size=12, color="#334155"), hovertemplate="<b>%{label}</b><br>%{value} operaciones<br>%{percent}<extra></extra>")])
    fig_par.update_layout(**chart_layout("Frecuencia de Pares", height=400))
    st.plotly_chart(fig_par, use_container_width=True)

colC, colD = st.columns(2)

with colC:
    resultado_counts = df_operaciones["Resultado"].value_counts().reindex(["Win", "Loss"]).fillna(0)
    fig_result = go.Figure(go.Bar(x=["Win", "Loss"], y=[resultado_counts.get("Win", 0), resultado_counts.get("Loss", 0)], marker=dict(color=["#2563eb", "#ef4444"], cornerradius=8, line=dict(width=0)), text=[f"{int(resultado_counts.get('Win', 0))}", f"{int(resultado_counts.get('Loss', 0))}"], textposition="outside", textfont=dict(family="Inter", size=14, color="#334155"), hovertemplate="<b>%{x}</b>: %{y}<extra></extra>"))
    fig_result.update_layout(**chart_layout("Win vs Loss", "Resultado", "Cantidad"))
    st.plotly_chart(fig_result, use_container_width=True)

with colD:
    tipo_count = df[df["Tipo"] != "Inicial"]["Tipo"].value_counts()
    fig_tipo = go.Figure(data=[go.Pie(labels=tipo_count.index, values=tipo_count.values, hole=0.55, marker=dict(colors=["#2563eb", "#0ea5e9"], line=dict(color="#fff", width=3)), textfont=dict(family="Inter", size=12, color="#334155"), hovertemplate="<b>%{label}</b><br>%{value}<br>%{percent}<extra></extra>")])
    fig_tipo.update_layout(**chart_layout("Compra vs Venta", height=400))
    st.plotly_chart(fig_tipo, use_container_width=True)

df_sin_inicial = df[df["Tipo"] != "Inicial"].copy()
conteo_total = df_sin_inicial.groupby("Par")["Resultado"].count()
conteo_wins = df_sin_inicial[df_sin_inicial["Resultado"] == "Win"].groupby("Par")["Resultado"].count()
conteo_losses = df_sin_inicial[df_sin_inicial["Resultado"] == "Loss"].groupby("Par")["Resultado"].count()
df_wr = pd.DataFrame({"Total": conteo_total, "Wins": conteo_wins, "Losses": conteo_losses}).fillna(0)
df_wr["% Winrate"] = df_wr["Wins"] / df_wr["Total"] * 100
df_wr["% Lossrate"] = df_wr["Losses"] / df_wr["Total"] * 100
df_wr["Barra"] = df_wr["% Winrate"] - df_wr["% Lossrate"]
df_wr.reset_index(inplace=True)
fig_wr_par = go.Figure()
fig_wr_par.add_trace(go.Bar(x=df_wr["Par"], y=df_wr["Barra"], marker=dict(color=df_wr["Barra"].apply(lambda x: "#2563eb" if x >= 0 else "#ef4444"), cornerradius=6), text=df_wr["% Winrate"].apply(lambda x: f"{x:.1f}%"), textposition="outside", textfont=dict(family="Inter", size=12, color="#334155"), hovertemplate="<b>%{x}</b><br>Winrate neto: %{y:.1f}%<extra></extra>"))
fig_wr_par.update_layout(**chart_layout("Winrate por Par", "Par", "Diferencia % Wins - Losses"))
st.plotly_chart(fig_wr_par, use_container_width=True)

fig_pips = go.Figure()
fig_pips.add_trace(go.Box(y=df["SL (pips)"], name="Stop Loss", marker_color="#ef4444", fillcolor="rgba(239,68,68,0.1)", line=dict(color="#ef4444", width=2), boxmean='sd'))
fig_pips.add_trace(go.Box(y=df["TP (pips)"], name="Take Profit", marker_color="#2563eb", fillcolor="rgba(37,99,235,0.1)", line=dict(color="#2563eb", width=2), boxmean='sd'))
fig_pips.update_layout(**chart_layout("Distribución SL vs TP", "", "Pips", showlegend=True))
st.plotly_chart(fig_pips, use_container_width=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║              RENDIMIENTO POR PERIODO (CORREGIDO V3)             ║
# ╚══════════════════════════════════════════════════════════════════╝
st.markdown('<div class="section-title" style="animation-delay:0.5s;">Rendimiento por Período</div>', unsafe_allow_html=True)
st.markdown('<div class="section-sub">Análisis porcentual diario, semanal y mensual</div>', unsafe_allow_html=True)

df_rend_base = df[df["Tipo"] != "Inicial"].copy()
balance_ref = st.session_state.balance_inicial
tiene_balance = balance_ref > 0
y_unit = "%" if tiene_balance else "USD"

if not tiene_balance and not df_rend_base.empty:
    st.markdown('<div class="balance-warning">⚠️ Ingresa tu <b>Balance TOTAL de la cuenta</b> en el sidebar para ver porcentajes reales. Actualmente se muestra en USD.</div>', unsafe_allow_html=True)

def calcular_pct_progresivo(df_agrupado, col_resultado, balance_base):
    bal_acum = 0
    pcts = []
    for _, row in df_agrupado.iterrows():
        balance_periodo = balance_base + bal_acum
        if balance_periodo > 0:
            pct = (row[col_resultado] / balance_periodo) * 100
        else:
            pct = 0
        pcts.append(round(pct, 2))
        bal_acum += row[col_resultado]
    return pcts

# DIARIO
df_rend_base["Día"] = df_rend_base["Fecha"].dt.date
rend_diario = df_rend_base.groupby("Día")["Resultado USD"].sum().reset_index()
if tiene_balance:
    rend_diario["valor"] = calcular_pct_progresivo(rend_diario, "Resultado USD", balance_ref)
else:
    rend_diario["valor"] = rend_diario["Resultado USD"].round(2)

# SEMANAL
df_rend_base["Semana_start"] = df_rend_base["Fecha"].dt.to_period("W").apply(lambda p: p.start_time)
rend_semanal = df_rend_base.groupby("Semana_start")["Resultado USD"].sum().reset_index()
if tiene_balance:
    rend_semanal["valor"] = calcular_pct_progresivo(rend_semanal, "Resultado USD", balance_ref)
else:
    rend_semanal["valor"] = rend_semanal["Resultado USD"].round(2)

# MENSUAL
df_rend_base["Mes_start"] = df_rend_base["Fecha"].dt.to_period("M").apply(lambda p: p.start_time)
rend_mensual = df_rend_base.groupby("Mes_start")["Resultado USD"].sum().reset_index()
if tiene_balance:
    rend_mensual["valor"] = calcular_pct_progresivo(rend_mensual, "Resultado USD", balance_ref)
else:
    rend_mensual["valor"] = rend_mensual["Resultado USD"].round(2)

for df_rend in [rend_diario, rend_semanal, rend_mensual]:
    df_rend["valor"].replace([np.inf, -np.inf], np.nan, inplace=True)
    df_rend["valor"].fillna(0, inplace=True)

if tiene_balance:
    text_diario = [f"{v:+.1f}%" for v in rend_diario["valor"]]
    text_semanal = [f"{v:+.1f}%" for v in rend_semanal["valor"]]
    text_mensual = [f"{v:+.1f}%" for v in rend_mensual["valor"]]
else:
    text_diario = [f"${v:+,.0f}" for v in rend_diario["valor"]]
    text_semanal = [f"${v:+,.0f}" for v in rend_semanal["valor"]]
    text_mensual = [f"${v:+,.0f}" for v in rend_mensual["valor"]]

colors_diario = ["#2563eb" if v >= 0 else "#ef4444" for v in rend_diario["valor"]]
colors_semanal = ["#2563eb" if v >= 0 else "#ef4444" for v in rend_semanal["valor"]]
colors_mensual = ["#2563eb" if v >= 0 else "#ef4444" for v in rend_mensual["valor"]]

x_diario = [d.strftime("%d %b") if hasattr(d, 'strftime') else str(d) for d in rend_diario["Día"]]
x_semanal = [d.strftime("%d %b %Y") for d in rend_semanal["Semana_start"]]
x_mensual = [d.strftime("%b %Y") for d in rend_mensual["Mes_start"]]

col_g1, col_g2, col_g3 = st.columns(3)

with col_g1:
    fig_d = go.Figure()
    fig_d.add_trace(go.Bar(x=x_diario, y=rend_diario["valor"].tolist(), marker=dict(color=colors_diario, cornerradius=4), text=text_diario, textposition="outside", textfont=dict(family="Inter", size=10, color="#64748b"), hovertemplate="<b>%{x}</b><br>%{text}<extra></extra>"))
    fig_d.update_layout(**chart_layout("Diario", "Día", y_unit, height=350))
    fig_d.update_layout(xaxis=dict(tickangle=-45))
    st.plotly_chart(fig_d, use_container_width=True)

with col_g2:
    fig_s = go.Figure()
    fig_s.add_trace(go.Bar(x=x_semanal, y=rend_semanal["valor"].tolist(), marker=dict(color=colors_semanal, cornerradius=4), text=text_semanal, textposition="outside", textfont=dict(family="Inter", size=10, color="#64748b"), hovertemplate="<b>%{x}</b><br>%{text}<extra></extra>"))
    fig_s.update_layout(**chart_layout("Semanal", "Semana", y_unit, height=350))
    fig_s.update_layout(xaxis=dict(tickangle=-45))
    st.plotly_chart(fig_s, use_container_width=True)

with col_g3:
    fig_m = go.Figure()
    fig_m.add_trace(go.Bar(x=x_mensual, y=rend_mensual["valor"].tolist(), marker=dict(color=colors_mensual, cornerradius=4), text=text_mensual, textposition="outside", textfont=dict(family="Inter", size=10, color="#64748b"), hovertemplate="<b>%{x}</b><br>%{text}<extra></extra>"))
    fig_m.update_layout(**chart_layout("Mensual", "Mes", y_unit, height=350))
    fig_m.update_layout(xaxis=dict(tickangle=-45))
    st.plotly_chart(fig_m, use_container_width=True)


st.markdown('<div class="section-title">Indicadores de Pips</div>', unsafe_allow_html=True)
st.markdown('<div class="section-sub">Promedio de pips ganados y perdidos por operación</div>', unsafe_allow_html=True)
ops_validas = df_operaciones.copy()
pips_ganados = ops_validas[ops_validas["Resultado USD"] > 0]["TP (pips)"]
pips_perdidos = ops_validas[ops_validas["Resultado USD"] < 0]["SL (pips)"]
tp_promedio_ganador = pips_ganados.mean() if not pips_ganados.empty else 0
sl_promedio_perdedor = pips_perdidos.mean() if not pips_perdidos.empty else 0

st.markdown(f"""
<div class="kpi-grid" style="grid-template-columns: repeat(2, 1fr);">
    <div class="kpi-card"><div class="kpi-icon green">🎯</div><div class="kpi-label">TP promedio cuando ganas</div><div class="kpi-value success">{tp_promedio_ganador:.1f} pips</div><div class="kpi-sub">Promedio de Take Profit ejecutado</div></div>
    <div class="kpi-card"><div class="kpi-icon red">🛑</div><div class="kpi-label">SL promedio cuando pierdes</div><div class="kpi-value danger">{sl_promedio_perdedor:.1f} pips</div><div class="kpi-sub">Promedio de Stop Loss ejecutado</div></div>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="section-title">Tipo de Operación más Rentable</div>', unsafe_allow_html=True)
st.markdown('<div class="section-sub">¿Con qué tipo te va mejor?</div>', unsafe_allow_html=True)
df_tipo = df_operaciones.copy()
ganancias_por_tipo = df_tipo.groupby("Tipo")["Resultado USD"].sum()
tipo_mas_rentable = ganancias_por_tipo.idxmax() if not ganancias_por_tipo.empty else "N/A"
ganancia_mejor_tipo = ganancias_por_tipo.max() if not ganancias_por_tipo.empty else 0

st.markdown(f"""
<div class="kpi-grid" style="grid-template-columns: repeat(2, 1fr);">
    <div class="kpi-card"><div class="kpi-icon blue">📌</div><div class="kpi-label">Tipo más rentable</div><div class="kpi-value accent">{tipo_mas_rentable}</div><div class="kpi-sub">Basado en ganancia neta acumulada</div></div>
    <div class="kpi-card"><div class="kpi-icon green">💰</div><div class="kpi-label">Ganancia neta con ese tipo</div><div class="kpi-value success">${ganancia_mejor_tipo:,.2f}</div><div class="kpi-sub">Total acumulado en USD</div></div>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="section-title">Modelo Predictivo</div>', unsafe_allow_html=True)
st.markdown('<div class="section-sub">Tendencia de rentabilidad en el tiempo con regresión lineal</div>', unsafe_allow_html=True)

df_pred = df_operaciones.copy()
df_pred['Fecha_Ordinal'] = df_pred['Fecha'].map(datetime.toordinal)
df_pred_grouped = df_pred.groupby('Fecha_Ordinal')['Resultado USD'].sum().reset_index()
df_pred_grouped['Fecha'] = df_pred_grouped['Fecha_Ordinal'].map(datetime.fromordinal)

from sklearn.linear_model import LinearRegression
X = df_pred_grouped[['Fecha_Ordinal']]
y = df_pred_grouped['Resultado USD'].cumsum()
model = LinearRegression()
model.fit(X, y)
df_pred_grouped['Real'] = y
df_pred_grouped['Predicha'] = model.predict(X)

fig_pred = go.Figure()
fig_pred.add_trace(go.Scatter(x=df_pred_grouped['Fecha'], y=df_pred_grouped['Real'], mode='lines+markers', name='Rentabilidad Real', line=dict(color='#2563eb', width=3, shape='spline'), marker=dict(size=6, color='#2563eb', line=dict(width=2, color='#fff')), fill='tozeroy', fillcolor='rgba(37, 99, 235, 0.05)', hovertemplate="<b>%{x|%Y-%m-%d}</b><br>Real: $%{y:,.0f}<extra></extra>"))
fig_pred.add_trace(go.Scatter(x=df_pred_grouped['Fecha'], y=df_pred_grouped['Predicha'], mode='lines', name='Tendencia Predicha', line=dict(color='#0ea5e9', width=2, dash='dash'), hovertemplate="<b>%{x|%Y-%m-%d}</b><br>Predicha: $%{y:,.0f}<extra></extra>"))
fig_pred.update_layout(**chart_layout("Rentabilidad Real vs Predicha", "Fecha", "USD", showlegend=True))
st.plotly_chart(fig_pred, use_container_width=True)

pendiente = model.coef_[0]
pendiente_color = "success" if pendiente >= 0 else "danger"
pendiente_txt = "↑ Mejora" if pendiente >= 0 else "↓ Deterioro"

st.markdown(f"""
<div class="kpi-grid" style="grid-template-columns: repeat(2, 1fr); margin-top: 16px;">
    <div class="kpi-card"><div class="kpi-icon {'green' if pendiente >= 0 else 'red'}">📐</div><div class="kpi-label">Pendiente de predicción</div><div class="kpi-value {pendiente_color}">{pendiente:.2f} USD/día</div><div class="kpi-sub">{pendiente_txt} — tendencia {'positiva' if pendiente >= 0 else 'negativa'}</div></div>
    <div class="kpi-card"><div class="kpi-icon blue">🧠</div><div class="kpi-label">Interpretación</div><div class="kpi-value" style="font-size:16px;">{'Tu estrategia mejora con el tiempo' if pendiente >= 0 else 'Revisa tu estrategia'}</div><div class="kpi-sub">Basado en regresión lineal de resultados</div></div>
</div>
""", unsafe_allow_html=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║              ANÁLISIS DE VALORES ATÍPICOS (OUTLIERS)            ║
# ╚══════════════════════════════════════════════════════════════════╝
st.markdown('<div class="section-title">Detección de Valores Atípicos</div>', unsafe_allow_html=True)
st.markdown('<div class="section-sub">Operaciones con resultado inusual respecto a tu patrón normal (método IQR)</div>', unsafe_allow_html=True)

if len(df_operaciones) >= 4:
    resultados = df_operaciones["Resultado USD"].values
    Q1 = np.percentile(resultados, 25)
    Q3 = np.percentile(resultados, 75)
    IQR = Q3 - Q1
    limite_inf = Q1 - 1.5 * IQR
    limite_sup = Q3 + 1.5 * IQR

    df_outliers = df_operaciones.copy()
    df_outliers["Es Atípico"] = (df_outliers["Resultado USD"] < limite_inf) | (df_outliers["Resultado USD"] > limite_sup)
    df_outliers["Op #"] = range(1, len(df_outliers) + 1)

    n_atipicos = df_outliers["Es Atípico"].sum()
    atipicos_vals = df_outliers[df_outliers["Es Atípico"]]["Resultado USD"]
    impacto_atipicos = atipicos_vals.sum() if not atipicos_vals.empty else 0
    impacto_sin_atipicos = ganancia_neta - impacto_atipicos
    mayor_atipico = atipicos_vals.abs().max() if not atipicos_vals.empty else 0
    pct_impacto = (impacto_atipicos / ganancia_neta * 100) if ganancia_neta != 0 else 0

    # Colores por punto
    colores = []
    for _, row in df_outliers.iterrows():
        if row["Es Atípico"] and row["Resultado USD"] > 0:
            colores.append("#f59e0b")  # Atípico positivo = amber
        elif row["Es Atípico"] and row["Resultado USD"] < 0:
            colores.append("#ef4444")  # Atípico negativo = rojo
        elif row["Resultado USD"] > 0:
            colores.append("#2563eb")  # Normal positivo
        else:
            colores.append("#94a3b8")  # Normal negativo

    fig_outlier = go.Figure()

    # Zona normal (banda IQR)
    fig_outlier.add_hrect(y0=limite_inf, y1=limite_sup, fillcolor="rgba(37, 99, 235, 0.04)", line_width=0, annotation_text="Rango normal", annotation_position="top left", annotation_font=dict(family="Inter", size=10, color="#94a3b8"))

    # Líneas de límite
    fig_outlier.add_hline(y=limite_sup, line_dash="dash", line_color="#f59e0b", line_width=1.5, annotation_text=f"Límite sup: ${limite_sup:+,.0f}", annotation_position="top right", annotation_font=dict(family="Inter", size=10, color="#f59e0b"))
    fig_outlier.add_hline(y=limite_inf, line_dash="dash", line_color="#ef4444", line_width=1.5, annotation_text=f"Límite inf: ${limite_inf:+,.0f}", annotation_position="bottom right", annotation_font=dict(family="Inter", size=10, color="#ef4444"))
    fig_outlier.add_hline(y=0, line_dash="dot", line_color="#cbd5e1", line_width=1)

    # Puntos
    fig_outlier.add_trace(go.Scatter(
        x=df_outliers["Op #"], y=df_outliers["Resultado USD"],
        mode="markers",
        marker=dict(
            size=df_outliers["Es Atípico"].apply(lambda x: 14 if x else 8).tolist(),
            color=colores,
            line=dict(width=2, color="#ffffff"),
            symbol=df_outliers["Es Atípico"].apply(lambda x: "diamond" if x else "circle").tolist()
        ),
        text=df_outliers.apply(lambda r: f"{'⚠️ ATÍPICO' if r['Es Atípico'] else 'Normal'}<br>{r['Par']} · {r['Tipo']}<br>${r['Resultado USD']:+,.0f}", axis=1),
        hovertemplate="%{text}<extra></extra>"
    ))

    fig_outlier.update_layout(**chart_layout("Mapa de Operaciones — Detección de Outliers", "Operación N°", "Resultado USD", height=420))
    fig_outlier.update_layout(xaxis=dict(tickmode='linear', tick0=1, dtick=1))
    st.plotly_chart(fig_outlier, use_container_width=True)

    # KPIs de outliers
    atipico_icon = "amber" if n_atipicos > 0 else "green"
    impacto_color = "danger" if impacto_atipicos < 0 else ("success" if impacto_atipicos > 0 else "accent")

    st.markdown(f"""
    <div class="kpi-grid">
        <div class="kpi-card"><div class="kpi-icon {atipico_icon}">⚠️</div><div class="kpi-label">Operaciones atípicas</div><div class="kpi-value">{n_atipicos} de {len(df_operaciones)}</div><div class="kpi-sub">Fuera del rango IQR ×1.5</div></div>
        <div class="kpi-card"><div class="kpi-icon {'red' if impacto_atipicos < 0 else 'green'}">💥</div><div class="kpi-label">Impacto de atípicos</div><div class="kpi-value {impacto_color}">${impacto_atipicos:+,.0f}</div><div class="kpi-sub">{abs(pct_impacto):.0f}% del P&L total</div></div>
        <div class="kpi-card"><div class="kpi-icon blue">📊</div><div class="kpi-label">P&L sin atípicos</div><div class="kpi-value accent">${impacto_sin_atipicos:+,.0f}</div><div class="kpi-sub">Tu resultado "real" sin extremos</div></div>
        <div class="kpi-card"><div class="kpi-icon red">🎯</div><div class="kpi-label">Mayor valor atípico</div><div class="kpi-value danger">${mayor_atipico:,.0f}</div><div class="kpi-sub">Rango normal: ${limite_inf:+,.0f} a ${limite_sup:+,.0f}</div></div>
    </div>
    """, unsafe_allow_html=True)

    # Interpretación automática
    if n_atipicos == 0:
        interp_html = """<div class="estado-badge rentable">✅ Sin operaciones atípicas — Tu trading es consistente, sin picos anormales</div>"""
    elif n_atipicos > 0 and impacto_atipicos > 0:
        interp_html = f"""<div class="estado-badge warning">⚠️ {n_atipicos} atípico(s) positivo(s) — Tus ganancias dependen de golpes de suerte. Sin esos trades tu P&L sería ${impacto_sin_atipicos:+,.0f}</div>"""
    elif n_atipicos > 0 and impacto_atipicos < 0:
        interp_html = f"""<div class="estado-badge danger">❌ {n_atipicos} atípico(s) negativo(s) — Pérdidas extremas dañan tu cuenta. Sin esos trades tu P&L sería ${impacto_sin_atipicos:+,.0f}</div>"""
    else:
        interp_html = f"""<div class="estado-badge neutral">🔍 {n_atipicos} atípico(s) detectado(s) — Revisa si son errores de gestión o eventos extraordinarios</div>"""
    st.markdown(interp_html, unsafe_allow_html=True)

else:
    st.markdown("""
    <div class="no-news-card">
        <div class="no-news-icon">📊</div>
        <div class="no-news-title">Se necesitan al menos 4 operaciones</div>
        <div class="no-news-sub">Agrega más trades para que el análisis de valores atípicos sea significativo.</div>
    </div>
    """, unsafe_allow_html=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║                     NOTICIAS                                    ║
# ╚══════════════════════════════════════════════════════════════════╝
st.markdown('<div class="section-title">Noticias de Alto Impacto</div>', unsafe_allow_html=True)
st.markdown('<div class="section-sub">Calendario económico — eventos de alto impacto hoy</div>', unsafe_allow_html=True)

def obtener_noticias_alto_impacto():
    try:
        url = "https://www.investing.com/economic-calendar/"
        headers = {"User-Agent": "Mozilla/5.0"}
        response = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.text, "html.parser")
        noticias = []
        for fila in soup.select("tr.js-event-item"):
            impacto = fila.select_one(".sentiment")
            if impacto and len(impacto.select(".grayFullBullishIcon")) >= 3:
                hora = fila.select_one(".first.left.time") or fila.select_one(".time")
                moneda = fila.select_one(".left.flagCur.noWrap")
                evento = fila.select_one(".event")
                noticias.append({"hora": hora.text.strip() if hora else "", "moneda": moneda.text.strip() if moneda else "", "evento": evento.text.strip() if evento else ""})
        return noticias
    except Exception as e:
        return f"Error al obtener noticias: {e}"

noticias_hoy = obtener_noticias_alto_impacto()
if isinstance(noticias_hoy, str):
    st.markdown(f"""
    <div class="no-news-card">
        <div class="no-news-icon">⚠️</div>
        <div class="no-news-title">Error al cargar noticias</div>
        <div class="no-news-sub">{noticias_hoy}</div>
    </div>
    """, unsafe_allow_html=True)
elif noticias_hoy:
    for i, noticia in enumerate(noticias_hoy):
        st.markdown(f"""
        <div class="news-item" style="animation-delay:{0.05*i}s;">
            <div style="display:flex; align-items:center; gap:16px;">
                <div style="font-family:'JetBrains Mono',monospace; font-size:13px; color:var(--blue-500); font-weight:600; min-width:60px;">{noticia['hora']}</div>
                <div style="background:var(--blue-50); padding:4px 10px; border-radius:6px; font-size:11px; font-weight:700; color:var(--blue-600); letter-spacing:0.05em;">{noticia['moneda']}</div>
                <div style="font-size:13px; color:var(--text); font-weight:600;">{noticia['evento']}</div>
                <div style="margin-left:auto; width:8px; height:8px; border-radius:50%; background:#ef4444; box-shadow:0 0 6px rgba(239,68,68,0.4);"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)
else:
    st.markdown("""
    <div class="no-news-card">
        <div class="no-news-icon">📭</div>
        <div class="no-news-title">Sin noticias de alto impacto hoy</div>
        <div class="no-news-sub">No se encontraron eventos económicos de alto impacto programados para hoy. Buen momento para operar con tranquilidad.</div>
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")

# ╔══════════════════════════════════════════════════════════════════╗
# ║             OPERACIONES REGISTRADAS                             ║
# ╚══════════════════════════════════════════════════════════════════╝
with st.expander("📋  Ver todas las operaciones registradas"):
    st.dataframe(df, use_container_width=True, hide_index=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║                  DESCARGA CSV + EXCEL                           ║
# ╚══════════════════════════════════════════════════════════════════╝
col_dl1, col_dl2 = st.columns(2)

with col_dl1:
    st.download_button(
        label="⬇️  Descargar CSV actualizado",
        data=df.to_csv(index=False).encode("utf-8"),
        file_name="operaciones_trading_actualizado.csv",
        mime="text/csv"
    )

with col_dl2:
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        def generar_excel(df_export, balance_ini, balance_fin, total_ops, winrate, ganancia_neta, profit_factor, max_dd, roi):
            wb = openpyxl.Workbook()
            ws_resumen = wb.active
            ws_resumen.title = "Resumen"

            header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1a4f8b", end_color="1a4f8b", fill_type="solid")
            kpi_font = Font(name="Calibri", size=11, bold=True)
            val_font = Font(name="Calibri", size=11)
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )

            ws_resumen.merge_cells("A1:C1")
            ws_resumen["A1"] = "TRADING DASHBOARD — RESUMEN"
            ws_resumen["A1"].font = Font(name="Calibri", size=16, bold=True, color="1a4f8b")
            ws_resumen["A1"].alignment = Alignment(horizontal="center")

            ws_resumen.merge_cells("A2:C2")
            ws_resumen["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ws_resumen["A2"].font = Font(name="Calibri", size=9, color="808080")
            ws_resumen["A2"].alignment = Alignment(horizontal="center")

            kpis = [
                ("Balance Inicial", f"${balance_ini:,.0f}"),
                ("Balance Final", f"${balance_fin:,.0f}"),
                ("Total Operaciones", f"{total_ops}"),
                ("Winrate", f"{winrate*100:.1f}%"),
                ("Ganancia Neta", f"${ganancia_neta:,.0f}"),
                ("Profit Factor", f"{profit_factor:.2f}" if profit_factor != float('inf') else "∞"),
                ("Máx Drawdown", f"${max_dd:,.0f}"),
                ("ROI", f"{roi:+.2f}%"),
            ]

            for row_idx, (kpi_name, kpi_val) in enumerate(kpis, start=4):
                cell_name = ws_resumen.cell(row=row_idx, column=1, value=kpi_name)
                cell_name.font = kpi_font
                cell_name.border = thin_border
                cell_val = ws_resumen.cell(row=row_idx, column=2, value=kpi_val)
                cell_val.font = val_font
                cell_val.alignment = Alignment(horizontal="right")
                cell_val.border = thin_border

            ws_resumen.column_dimensions["A"].width = 24
            ws_resumen.column_dimensions["B"].width = 20
            ws_resumen.column_dimensions["C"].width = 14

            ws_ops = wb.create_sheet("Operaciones")
            df_export_clean = df_export.copy()
            if "Fecha" in df_export_clean.columns:
                df_export_clean["Fecha"] = df_export_clean["Fecha"].astype(str)

            for r_idx, row in enumerate(dataframe_to_rows(df_export_clean, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_ops.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.font = Font(name="Calibri", size=10)

            for col_cells in ws_ops.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws_ops.column_dimensions[col_letter].width = min(max_len + 4, 30)

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

        excel_bytes = generar_excel(df, balance_inicial, balance_final, total_ops, winrate, ganancia_neta, profit_factor, max_dd, roi)
        st.download_button(
            label="📊  Descargar Excel con Resumen",
            data=excel_bytes,
            file_name="trading_dashboard_reporte.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.spreadsheetml"
        )
    except ImportError:
        st.warning("Instala openpyxl para exportar Excel: `pip install openpyxl`")

st.markdown("""
<div style='text-align:center; padding:40px 0 20px 0; color:#94a3b8; font-size:10px; font-family:Inter,sans-serif; letter-spacing:0.08em; font-weight:600; text-transform:uppercase;'>
    Trading Dashboard Pro · Built with Streamlit & Plotly · Designed for Traders
</div>
""", unsafe_allow_html=True)
